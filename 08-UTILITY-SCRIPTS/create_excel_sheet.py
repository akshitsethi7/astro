#!/usr/bin/env python3
"""
Create an Excel file with all astrological charts, color coding, and calculated strengths.
This is an alternative to Google Sheets that doesn't require API setup.
"""

import json
from datetime import datetime
from typing import Dict, List, Any
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")

# Import chart data and functions - we'll define them here to avoid import issues
# For now, we'll import from the main script, but handle ImportError gracefully
try:
    from create_google_sheet import (
        CHARTS_DATA, DETAILED_CHARTS, DASHA_DATA,
        get_planet_status, get_color_for_status, get_planet_name,
        EXALTATION, DEBILITATION, OWN_SIGNS, BENEFIC, MALEFIC
    )
except ImportError:
    # If import fails, we'll need to define these here
    print("Warning: Could not import from create_google_sheet.py")
    print("Please make sure create_google_sheet.py is in the same directory")
    CHARTS_DATA = {}
    DETAILED_CHARTS = {}
    DASHA_DATA = {}

def rgb_to_hex(rgb_dict: Dict[str, float]) -> str:
    """Convert RGB dict (0-1) to hex color."""
    r = int(rgb_dict.get("red", 1.0) * 255)
    g = int(rgb_dict.get("green", 1.0) * 255)
    b = int(rgb_dict.get("blue", 1.0) * 255)
    return f"{r:02X}{g:02X}{b:02X}"

def create_excel_file(filename: str = "Akshit_Birth_Chart_Analysis.xlsx"):
    """Create Excel file with all charts and features."""
    
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        return None
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Combine all charts
    all_charts = {**CHARTS_DATA, **DETAILED_CHARTS}
    
    # Create chart sheets
    sheet_order = ["D1", "D2", "D3", "D4", "D6", "D7", "D9", "D10", "D12", "D16", "D20", "D24", "D27", "D30", "D40", "D45", "D60"]
    
    for chart_id in sheet_order:
        if chart_id in all_charts:
            create_chart_sheet_excel(wb, chart_id, all_charts[chart_id])
    
    # Create dasha sheets
    create_dasha_sheet_excel(wb, "Mahadasha", DASHA_DATA["Mahadasha"])
    create_dasha_sheet_excel(wb, "Antardasha", DASHA_DATA["Antardasha"])
    create_dasha_sheet_excel(wb, "Pratyantardasha", DASHA_DATA["Pratyantardasha"])
    create_dasha_sheet_excel(wb, "Sookshmadasha", DASHA_DATA["Sookshmadasha"])
    
    # Create summary dashboard
    create_summary_dashboard_excel(wb, all_charts)
    
    # Save file
    wb.save(filename)
    print(f"✅ Excel file created: {filename}")
    return filename

def create_chart_sheet_excel(wb, chart_id: str, chart_data: Dict):
    """Create a sheet for a specific chart with formatting."""
    ws = wb.create_sheet(title=chart_id)
    
    # Chart title
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = chart_data["name"]
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    
    # Headers
    headers = ["Planet", "Sign", "Sign Lord", "Degree", "House", "Status", "Strength", "Benefic/Malefic", "Own Sign", "Exalted", "Debilitated", "Retrograde"]
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add planet data
    row_num = 4
    for planet_data in chart_data["planets"]:
        planet = planet_data["Planet"]
        sign = planet_data["Sign"]
        sign_lord = planet_data["Sign Lord"]
        degree = planet_data["Degree"]
        house = planet_data["House"]
        
        status = get_planet_status(planet, sign, sign_lord)
        
        # Status text
        status_parts = []
        if status["is_exalted"]:
            status_parts.append("Exalted")
        elif status["is_own_sign"]:
            status_parts.append("Own Sign")
        elif status["is_debilitated"]:
            status_parts.append("Debilitated")
        
        status_text = ", ".join(status_parts) if status_parts else "Normal"
        
        row = [
            planet,
            sign,
            sign_lord,
            degree,
            house,
            status_text,
            status["strength"],
            "Benefic" if status["is_benefic"] else "Malefic" if status["is_malefic"] else "Neutral",
            "Yes" if status["is_own_sign"] else "No",
            "Yes" if status["is_exalted"] else "No",
            "Yes" if status["is_debilitated"] else "No",
            "Yes" if status["is_retrograde"] else "No"
        ]
        
        ws.append(row)
        
        # Apply color formatting
        color = get_color_for_status(status)
        hex_color = rgb_to_hex(color)
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
        
        row_num += 1
    
    # Auto-adjust column widths
    column_widths = [12, 12, 12, 10, 8, 15, 10, 15, 10, 10, 12, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add auto-filter
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{row_num - 1}"

def create_dasha_sheet_excel(wb, sheet_name: str, dasha_data: List[Dict]):
    """Create a sheet for dasha periods."""
    ws = wb.create_sheet(title=sheet_name)
    
    headers = ["Planet/Period", "Start Date", "End Date", "Duration"]
    ws.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for period in dasha_data:
        start = period["Start Date"]
        end = period["End Date"]
        
        # Calculate duration
        duration = "N/A"
        if start != "Birth" and "-" in start and "-" in end:
            try:
                start_dt = datetime.strptime(start, "%d-%b-%Y %H:%M")
                end_dt = datetime.strptime(end, "%d-%b-%Y %H:%M")
                delta = end_dt - start_dt
                years = delta.days / 365.25
                duration = f"{years:.2f} years"
            except:
                pass
        
        row = [period["Planet"], start, end, duration]
        ws.append(row)
    
    # Auto-adjust column widths
    column_widths = [20, 20, 20, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(dasha_data) + 1}"

def create_summary_dashboard_excel(wb, all_charts: Dict):
    """Create a summary dashboard sheet."""
    ws = wb.create_sheet(title="Dashboard")
    
    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "ASTROLOGICAL CHART SUMMARY DASHBOARD"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.fill = PatternFill(start_color="3366CC", end_color="3366CC", fill_type="solid")
    
    # Headers
    ws.append([])
    headers = ["Chart", "Total Planets", "Avg Strength", "Benefic Count", "Malefic Count"]
    ws.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for chart_id in sorted(all_charts.keys()):
        chart = all_charts[chart_id]
        planets = chart["planets"]
        
        total = len(planets)
        strengths = []
        benefic_count = 0
        malefic_count = 0
        
        for p in planets:
            status = get_planet_status(p["Planet"], p["Sign"], p["Sign Lord"])
            strengths.append(status["strength"])
            if status["is_benefic"]:
                benefic_count += 1
            elif status["is_malefic"]:
                malefic_count += 1
        
        avg_strength = sum(strengths) / len(strengths) if strengths else 0
        
        ws.append([chart_id, total, f"{avg_strength:.1f}", benefic_count, malefic_count])
    
    # Auto-adjust column widths
    column_widths = [10, 12, 12, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add auto-filter
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{len(all_charts) + 3}"

if __name__ == "__main__":
    print("Creating Excel file with all astrological charts...")
    print()
    
    filename = create_excel_file()
    
    if filename:
        print(f"\n✅ Success! Excel file created: {filename}")
        print("\nFeatures included:")
        print("  ✓ Multiple chart sheets (D1, D2, D3, D4, D6, D7, D9, D10, D12, D16, D20, D24, D27, D30, D40, D45, D60)")
        print("  ✓ Dasha sheets (Mahadasha, Antardasha, Pratyantardasha, Sookshmadasha)")
        print("  ✓ Color coding (benefic/malefic/own sign/exalted/debilitated)")
        print("  ✓ Auto-calculated strengths")
        print("  ✓ Filters on all sheets")
        print("  ✓ Summary dashboard")
        print("\nYou can open this file in Excel, Google Sheets, or any spreadsheet application.")
    else:
        print("\n❌ Failed to create Excel file.")

